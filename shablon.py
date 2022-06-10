import openpyxl
from openpyxl.styles.borders import Border, Side

thick_border = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))


def shablon(savesheet, sheetfile2):
    savesheet["B6"] = sheetfile2["B6"].value
    savesheet.merge_cells('B6:B7')
    savesheet["C6"] = sheetfile2["C6"].value
    savesheet.merge_cells('C6:C7')
    savesheet["D6"] = sheetfile2["D6"].value
    savesheet.merge_cells('D6:H6')
    savesheet["I6"] = sheetfile2["I6"].value
    savesheet.merge_cells('I6:K6')
    savesheet["L6"] = sheetfile2["L6"].value
    savesheet.merge_cells('L6:L7')
    savesheet["M6"] = sheetfile2["M6"].value
    savesheet.merge_cells('M6:M7')
    savesheet["N6"] = sheetfile2["N6"].value
    savesheet.merge_cells('N6:N7')
    savesheet["O6"] = sheetfile2["O6"].value
    savesheet.merge_cells('O6:O7')
    savesheet["P6"] = sheetfile2["P6"].value
    savesheet.merge_cells('P6:P7')
    savesheet["Q6"] = sheetfile2["Q6"].value
    savesheet.merge_cells('Q6:Q7')
    savesheet["D7"] = sheetfile2["D7"].value
    savesheet["E7"] = sheetfile2["E7"].value
    savesheet["F7"] = sheetfile2["F7"].value
    savesheet["G7"] = sheetfile2["G7"].value
    savesheet["H7"] = sheetfile2["H7"].value
    savesheet["I7"] = sheetfile2["I7"].value
    savesheet["J7"] = sheetfile2["J7"].value
    savesheet["K7"] = sheetfile2["K7"].value
    savesheet.row_dimensions[6].height = 60
    savesheet.row_dimensions[7].height = 60
    savesheet.row_dimensions[8].height = 20
    savesheet.column_dimensions['A'].width = 5
    savesheet.column_dimensions['B'].width = 20
    savesheet.column_dimensions['C'].width = 15
    savesheet.column_dimensions['D'].width = 8
    savesheet.column_dimensions['E'].width = 8
    savesheet.column_dimensions['F'].width = 8
    savesheet.column_dimensions['G'].width = 8
    savesheet.column_dimensions['H'].width = 10
    savesheet.column_dimensions['I'].width = 8
    savesheet.column_dimensions['J'].width = 8
    savesheet.column_dimensions['K'].width = 10
    savesheet.column_dimensions['L'].width = 15
    savesheet.column_dimensions['P'].width = 10

    try:
        for i in range(17):
            savesheet[8][i].value = str(i+1)
    except:
        pass
    for i in range(17):
        for j in range(50):
            savesheet[j+1][i].alignment = openpyxl.styles.Alignment(
                horizontal="center", vertical="center", wrap_text=True)
            savesheet[j+1][i].font = openpyxl.styles.Font(
                bold=False, color="000000", name='Times New Roman', size=10)


if __name__ == "__main__":
    shablon('file2.xlsx')
