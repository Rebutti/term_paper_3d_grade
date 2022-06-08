import openpyxl
from shablon import shablon
from rentabel import rentabel
from openpyxl import load_workbook


# def reader(pathfile1, pathfile2, ser_nav_nav, ser_zar_plat, ESV, pev_vel):
def findgname(sheet):
    groupname = sheet["A2"].value
    groupname = groupname.replace(' ', '')
    groupname = groupname[5:7]
    return groupname


def findgfname(sheet):
    groupname = sheet["A1"].value
    groupname = groupname.replace(' ', '')
    groupname = groupname[groupname.find('-')+1:]
    return groupname.replace('-', ' ')


def findgtype(sheet):
    groupname = sheet["A2"].value
    groupname = groupname[groupname.find('(')+1:]
    groupname = groupname[:groupname.find(' ')]
    return groupname


def toFixed(numObj, digits=0):
    return f"{numObj:.{digits}f}"


def allmoney(I9, F9, G9, J9, H9, K9):

    # groupname = float(sheet["F"+str(row)].value) * \
    #     float(sheet["I"+str(row)].value)+float(sheet["G" +
    #                                                  str(row)].value)*float(sheet["J"+str(row)].value)+float(sheet["H"+str(row)].value)*float(sheet["K"+str(row)].value)
    money = toFixed(float(I9)*float(F9)+float(G9)
                    * float(J9)+float(H9)*float(K9))
    return money


def reader(values: dict):
    # for k, v in values.items():
    #     print(k, v)
    ser_nav_nav = values[0]
    ser_zar_plat = values[1]
    ESV = values[2]
    pev_vel = values[3]
    filepath1 = values[4]
    filepath2 = values[5]
    file2 = openpyxl.open(filepath2,
                          read_only=True, data_only=True)  # открываем файл
    sheetfile2 = file2.active  # переходим к первому листу в файле
    # вызываем класс, который потом станет выходным файлом
    if values["filetrue"] == True:
        savebook = load_workbook(values["dop_file"])
    else:
        savebook = openpyxl.Workbook()
        savebook.create_sheet(title='Лист1', index=0)
        savebook.remove(savebook['Sheet'])
    savesheet = savebook.active
    shablon(savesheet, sheetfile2)
    file1 = openpyxl.open(filepath1,
                          read_only=True, data_only=True)  # открываем файл
    sheetsfile1 = file1.sheetnames  # запоминаю все листы в файле
    for sheet in sheetsfile1:
        sheetfile1 = file1[sheet]
        rentabel(sheetfile1, sheetfile2, savesheet, ser_nav_nav,
                 ser_zar_plat, ESV, pev_vel, sheet, values=values)
    if values["filetrue"] == True:
        savebook.save(values['dop_file'])
    else:
        savebook.save('рентабельність.xlsx')
    file1.close()
    file2.close()
