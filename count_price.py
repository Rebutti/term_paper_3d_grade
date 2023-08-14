import openpyxl
from navantaj import toFixed, navantaj, findlpl, find_vibir_disc, findgtype, find_atec_ex, findatect
from rentabel import findcourse
import math
import xlwings as xw

def split_filename(filename):
    return filename.split('.')[0]


def count_price_file_open(filepathes, values):
    result_message = 'Результати збережені у файли:\n'
    for filepath1 in filepathes:
        counts = []
        print(filepath1)
        file1 = openpyxl.open(filepath1,
                            read_only=True, data_only=True)  # открываем файл
        sheetsfile1 = file1.sheetnames  # запоминаю все листы в файле
        if len(sheetsfile1) != 2 and len(sheetsfile1) != 4 and len(sheetsfile1) != 1 and len(sheetsfile1) != 3:
            file1.close()
            return 'У Вашому файлі неправильна кількість планів. Має бути 2 для магістрів, 4 для бакалаврів'
        elif len(sheetsfile1) == 4:
            file_save = openpyxl.load_workbook('розрахунок вартості шаблон бакалавр.xlsx')
        elif len(sheetsfile1) == 1:
            file_save = openpyxl.load_workbook('розрахунок вартості шаблон бакалавр 1 рік.xlsx')
        elif len(sheetsfile1) == 3:
            file_save = openpyxl.load_workbook('розрахунок вартості шаблон бакалавр 3 рік.xlsx')
        else:
            file_save = openpyxl.load_workbook('розрахунок вартості шаблон.xlsx')
        sheet_names = file_save.sheetnames
        sheet_number = 0
        for sheet in sheetsfile1:
            sheetfile1 = file1[sheet]
            result = count_price(sheetfile1, values=values, file_save=file_save, sheet_name=sheet_names[sheet_number])
            print(sheet_names[sheet_number])
            sheet_number+=1
            counts.append(result)
        
        for sheet in sheet_names:
            sheetfile1 = file_save[sheet]
        new_filename = split_filename(filepath1)+' розрахунок вартості.xlsx'
        file_save.save(new_filename)
        file1.close()
        result_message = result_message + new_filename+'\n'
        file_save.close()
        file_save1 = xw.Book(new_filename)
        gradation_of_salary(file_save1, values)
        file_save1.save()
        file_save1.close()
    return result_message

def gradation_of_salary(file, values):
    sheet_name = 'загальна вартість'
    sheet = file.sheets[sheet_name]
    amount_of_budgets, price_of_budget, navantajenya = find_profit_by_budget(sheet)
    price_for_budget = sheet['H3'].value
    price_for_contract = sheet['I3'].value
    print(f"кількість студентів бюджету для рентабельності = {amount_of_budgets, price_of_budget}, HPP = {values['НПП_витрати']}")
    sheet_name = 'рен групи'
    sheet = file.sheets[sheet_name]
    sheet['A4'].value = float(values['НПП_витрати'])/100
    sheet['B4'].value = amount_of_budgets
    sheet['C4'].value = toFixed(price_of_budget, 2)
    sheet['D7'].value = price_for_budget
    cell = sheet.range('D7')
    cell.api.Borders.Weight = 2
    sheet['E7'].value = price_for_contract
    cell = sheet.range('E7')
    cell.api.Borders.Weight = 2
    build_gradation(sheet, amount_of_budgets)
    sheet = file.sheets['ціноутворення']
    build_summary_data(sheet, values, amount_of_budgets, price_of_budget)
    # print(values)

    return None

def get_collumn_by_index(column_index):
    if column_index > 26:
        integer = column_index % 26 
        remainder = column_index // 26 
        if remainder > 26:
            return None
        column_letter = chr(64 + integer)+chr(64 + remainder)
    column_letter = chr(64 + column_index)
    return column_letter

def build_summary_data(sheet, values, amount_of_budgets, price_of_budget):
    letter = 'B'
    number = 5
    npp_value = int(values['НПП_витрати'])
    sheet['B'+str(number)].value = npp_value/100
    values['НПП_витрати1'] = npp_value
    # npp_bills = npp_counter(values, navantajenya)
    # all_bills = bill_counter(values, npp_bills)
    sheet['C'+str(number)].value = toFixed(amount_of_budgets, 2)
    sheet['D'+str(number)].value = toFixed(price_of_budget, 2)
    fst_pstn_for_contract = 8
    fst_pstn_for_budget = 5
    c_flag = False
    for step_b in range(int(values['ціноутв_мін_б']), int(values['ціноутв_макс_б'])+1, int(values['бюдж_крок'])):
        sheet['G'+str(fst_pstn_for_budget)].value = step_b
        cell = sheet.range('G'+str(fst_pstn_for_budget))
        cell.api.Borders.Weight = 2
        for step_c in range(int(values['ціноутв_мін_к']), int(values['ціноутв_макс_к'])+1, int(values['контр_крок'])):
            letter = get_collumn_by_index(fst_pstn_for_contract)
            if c_flag == False:
                sheet[letter+'4'].value = step_c
            cell = sheet.range(letter+'4')
            cell.api.Borders.Weight = 2
            cell = sheet.range(letter+str(fst_pstn_for_budget))
            cell.api.Borders.Weight = 2
            cell.number_format = '0.00'
            # =($C$5*$C$3-$G5*$C$3)/H$4
            fst_pstn_for_contract +=1
            sheet[letter+str(fst_pstn_for_budget)].value = '=($C$5*$C$3-$G'+str(fst_pstn_for_budget)+'*$C$3)/'+letter+'$4'
        if amount_of_budgets-1 <= step_b:
            break
        c_flag = True
        fst_pstn_for_contract = 8
        fst_pstn_for_budget +=1

    # number +=1

def build_gradation(sheet, amount_of_budgets):
    letter = 'B'
    number = 7
    formula_first_part = '=ЕСЛИ(ОКРВВЕРХ(($D$4-B'
    formula_second_part = '*$D$7)/$E$7,1) <= 0, 0, ОКРВВЕРХ(($D$4-B'
    formula_last_part = '*$D$7)/$E$7,1))'

    # =ЕСЛИ(ОКРВВЕРХ(($D$4-B7*$D$7)/$E$7,1) <= 0, 0, ОКРВВЕРХ(($D$4-B7*$D$7)/$E$7,1))
    for student in range(0, amount_of_budgets+1):
        cell = sheet.range('B'+str(number))
        cell.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignCenter
        cell.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter
        cell.api.Borders.Weight = 2
        sheet['B'+str(number)].value = student
        sheet['C'+str(number)].value = formula_first_part+str(number)+formula_second_part+str(number)+formula_last_part
        # cell.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignCenter
        # cell.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter
        number+=1

def find_profit_by_budget(sheet):
    budget_price = sheet.range('H3').value
    letter = 'G'
    number = 3
    for cell1 in range(1,91):
        if sheet.range(letter+str(number)).value <= budget_price:
            return int(sheet.range('A'+str(number)).value), sheet.range(letter+str(number)).value, sheet.range('C'+str(number)).value 
        else:
            number+=1
    return 90, 50000




def amount_of_groups_counter(amount_of_students, D1coeff):
    if amount_of_students <= D1coeff:
        return 1
    else:
        return math.ceil(amount_of_students/D1coeff)


def amount_of_subgroups_counter(amount_of_students, E1coeff):
    if amount_of_students <= E1coeff:
        return 1
    else:
        return math.ceil(amount_of_students/E1coeff)

def amount_of_potoks_counter(amount_of_students, potoks):
    if amount_of_students <= potoks:
        return 1
    else:
        return math.ceil(amount_of_students/potoks)


def find_weeks_amount(sheetfile1):
    letter = 'O'
    for row in range(1, 51):
        if str(sheetfile1[letter+str(row)].value).lower().find("тижнів") != -1:
            return float(sheetfile1[letter+str(row)].value.strip().replace(',','.').split()[0]), float(sheetfile1['S'+str(row)].value.strip().replace(',','.').split()[0])
    return None, None


def find_current_consultations(sheetfile1):
    letter = 'B'
    for row in range(1, 51):
        if str(sheetfile1[letter+str(row)].value).lower().find("разом") != -1 and len(str(sheetfile1[letter+str(row)].value).strip()) == 5:
            return int(sheetfile1['M'+str(row)].value)
    return None


def find_kr_kp(sheetfile1, last_row):
    letter = "K"
    result_kr = 0
    result_kp = 0
    for row in range(1, 51):
        if str(sheetfile1[letter+str(row)].value).lower().find("11") != -1:
            first_row = row+1
    for row in range(first_row, last_row):
        if str(sheetfile1[letter+str(row)].value).lower().find("кр") != -1:
            result_kr += 1
        elif str(sheetfile1[letter+str(row)].value).lower().find("кп") != -1:
            result_kr += 1
    return result_kr, result_kp


def find_amount_of_students(sheetfile1):
    letter = 'A'
    number = 1
    for row in range(1, 51):
        if sheetfile1[letter+str(row)].value != None:
            if 'кількість' in str(sheetfile1[letter+str(row)].value).lower():
                number = row
    kil_stud = sheetfile1["A"+str(number)].value.split(' ')
    kil_stud_new = []
    for student in kil_stud:
        if student != '':
            kil_stud_new.append(student)
    ind = kil_stud_new.index("студентів")+1
    kil_stud = int(kil_stud_new[ind]) + int(kil_stud_new[ind+2])
    return kil_stud


def npp_counter(values, navantaj):
    result = float(navantaj)/float(values[0])*12*float(values[1])*float(
        values[2])+navantaj/float(values[0])*float(values[3])
    return toFixed(result, 2)


def bill_counter(values, npp_bills):
    return toFixed(npp_bills/(float(values['НПП_витрати1']))*100, 2)

def find_vir_pr(sheetfile1, start_row):
    letter = 'B'
    kil_tij = 0
    for row in range(start_row, start_row+30):
        if sheetfile1[letter+str(row)].value != None:
            if str(sheetfile1[letter+str(row)].value).lower().find("виробнича:") != -1:
                kil_tij += float(sheetfile1['E'+str(row)].value)
    return kil_tij

def find_kval_rob(sheetfile1, start_row):
    letter = 'B'
    kil_rob = 0
    for row in range(start_row, start_row+30):
        if sheetfile1[letter+str(row)].value != None:
            if str(sheetfile1[letter+str(row)].value).lower().find("кваліфікаційна робота") != -1:
                # kil_rob += 1
                return 1
    return kil_rob

def get_price_for_budget(sheet_name, values):
    if sheet_name[0] == '1':
        return values['бюджет1']
    elif sheet_name[0] == '2':
        return values['бюджет2']
    elif sheet_name[0] == '3':
        return values['бюджет3']
    elif sheet_name[0] == '4':
        return values['бюджет4']
    elif sheet_name[0] == '5':
        return values['бюджет5']
    else:
        return values['бюджет6']
    
def get_price_for_contract(sheet_name, values):
    if sheet_name[0] == '1':
        return values['контракт1']
    elif sheet_name[0] == '2':
        return values['контракт2']
    elif sheet_name[0] == '3':
        return values['контракт3']
    elif sheet_name[0] == '4':
        return values['контракт4']
    elif sheet_name[0] == '5':
        return values['контракт5']
    else:
        return values['контракт6']


def count_price(sheetfile1, values, file_save, sheet_name):
    all_counts = []
    number1 = findlpl(sheetfile1)
    number = number1[0]
    kil_tij_1_cem, kil_tij_2_cem = find_weeks_amount(sheetfile1)
    if kil_tij_1_cem == None:
        return f'Не вдалось знайти кількість навчальних тижнів {sheetfile1.title}'
    lekciya1 = toFixed(float(
        sheetfile1["O"+str(number)].value if sheetfile1["O"+str(number)].value != None else 0), 2)
    praktika1 = toFixed(float(
        sheetfile1["P"+str(number)].value if sheetfile1["P"+str(number)].value != None else 0), 2)
    labi1 = toFixed(float(
        sheetfile1["Q"+str(number)].value if sheetfile1["Q"+str(number)].value != None else 0), 2)
    lekciya2 = toFixed(float(
        sheetfile1["S"+str(number)].value if sheetfile1["S"+str(number)].value != None else 0), 2)
    praktika2 = toFixed(float(
        sheetfile1["T"+str(number)].value if sheetfile1["T"+str(number)].value != None else 0), 2)
    labi2 = toFixed(float(
        sheetfile1["U"+str(number)].value if sheetfile1["U"+str(number)].value != None else 0), 2)
    exzamens = str(sheetfile1["G"+str(number)
                                      ].value).strip().split(' ')
    zaliki = str(sheetfile1["H"+str(number)].value).strip().split(' ')
    for el in zaliki:
        if el == '':
            zaliki.remove(el)
    for el in exzamens:
        if el == '':
            exzamens.remove(el)
    exzamens = [int(i) for i in exzamens]
    exzamens = sum(exzamens)
    zaliki = [int(i) for i in zaliki]
    zaliki = sum(zaliki)
    current_consultations = find_current_consultations(sheetfile1)
    if current_consultations == None:
        return f'Не вдалось знайти поточні консультації {sheetfile1.title}'
    result_kr, result_kp = find_kr_kp(sheetfile1, number)
    vibir_disc = 0
    if number1[1] == 1:
        vibir_disc = find_vibir_disc(sheetfile1, number)

    kil_atec_ex = find_atec_ex(sheetfile1,number)

    kil_tij_vir_pr = find_vir_pr(sheetfile1, number)
    kil_kval_rob = find_kval_rob(sheetfile1, number)
    row_number = 4
    kil_hours = toFixed(float(
        sheetfile1["M"+str(number)].value if sheetfile1["O"+str(number)].value != None else 0), 2)
    
    k_pot_kons = findgtype(sheetfile1)
    if k_pot_kons == "денна":
        k_pot_kons = values['пот_конс_денна']
        k_indiv = values['індивід_денна/вечірня']
        if findcourse(sheetfile1.title)[1] == 0:
            k_vibirkovi_disc = values['вибір_дисц_бакалавр_денна']
        else:
            k_vibirkovi_disc = values['вибір_дисц_магістр_денна']
    elif k_pot_kons == "вечірня":
        k_pot_kons = values['пот_конс_вечірня']
        k_indiv = values['індивід_денна/вечірня']
        k_vibirkovi_disc = values['вибір_дисц_бакалавр_вечірня']
    elif k_pot_kons == "заочна":
        k_pot_kons = values['пот_конс_заочна']
        k_indiv = values['індивід_заочна']
        if findcourse(sheetfile1.title)[1] == 0:
            k_vibirkovi_disc = values['вибір_дисц_бакалавр_заочна']
        else:
            k_vibirkovi_disc = values['вибір_дисц_магістр_заочна']
    elif k_pot_kons == "дуальна":
        k_pot_kons = values['пот_конс_дуальна']
        k_indiv = values['індивід_дуальна']
    else:
        k_pot_kons = 1

    file_save_sheet = file_save[sheet_name]
    if values['separate_budget'] == True:
        print('tytb')
        values['бюджет'] = get_price_for_budget(sheet_name, values)
    if values['separate_contract'] == True:
        print('tytc')
        values['контракт'] = get_price_for_contract(sheet_name, values)
    file_save_sheet['AB2'] = float(values['бюджет'])
    file_save_sheet['AC2'] = float(values['контракт'])

    kil_indiv_zavd = 0

    for student_number in range(1, 91):
        # for sheet_name in sheet_names:
        file_save_sheet['A'+str(row_number)] = student_number
        amount_of_potoks = amount_of_potoks_counter(
            student_number, int(values['поток']))
        file_save_sheet['C'+str(row_number)] = amount_of_potoks
        amount_of_groups = amount_of_groups_counter(
            student_number, int(values['D1coeff']))
        file_save_sheet['D'+str(row_number)] = amount_of_groups
        amount_of_subgroups = amount_of_subgroups_counter(
            student_number, int(values['E1coeff']))
        file_save_sheet['E'+str(row_number)] = amount_of_subgroups
        file_save_sheet['F'+str(row_number)] = kil_tij_1_cem
        file_save_sheet['G'+str(row_number)] = kil_tij_2_cem
        file_save_sheet['H'+str(row_number)] = lekciya1
        file_save_sheet['I'+str(row_number)] = praktika1
        file_save_sheet['J'+str(row_number)] = labi1
        file_save_sheet['K'+str(row_number)] = lekciya2
        file_save_sheet['L'+str(row_number)] = praktika2
        file_save_sheet['M'+str(row_number)] = labi2
        file_save_sheet['N'+str(row_number)] = exzamens
        file_save_sheet['O'+str(row_number)] = zaliki
        file_save_sheet['P'+str(row_number)] = toFixed(current_consultations, 2)
        file_save_sheet['Q'+str(row_number)] = toFixed(kil_indiv_zavd, 2)
        file_save_sheet['R'+str(row_number)] = result_kr
        file_save_sheet['S'+str(row_number)] = result_kp
        file_save_sheet['T'+str(row_number)] = kil_tij_vir_pr
        file_save_sheet['U'+str(row_number)] = 0
        file_save_sheet['V'+str(row_number)
                        ] = vibir_disc
        file_save_sheet['W'+str(row_number)] = kil_atec_ex
        file_save_sheet['X'+str(row_number)] = kil_kval_rob
        file_save_sheet['Y'+str(row_number)] = 0
        # kil_stud = find_amount_of_students(sheetfile1)
        # navantajenya2 = navantaj(
        #     sheetfile1, values, student_number, sheetfile1.title, findcourse(sheetfile1.title)[1], 0)
        
        navantajenya = float(lekciya1)*float(kil_tij_1_cem)*amount_of_potoks+float(praktika1)*float(kil_tij_1_cem)*float(amount_of_groups)+float(labi1)*float(kil_tij_1_cem)*float(amount_of_subgroups)+float(lekciya2)*float(kil_tij_2_cem)*float(amount_of_potoks)+float(praktika2)*float(kil_tij_2_cem)*float(amount_of_groups)+float(labi2)*float(kil_tij_2_cem)*float(amount_of_subgroups)+float(exzamens)*float(values['пров_екз'])+float(exzamens)*float(values['конс_пред_екз'])*float(zaliki)*float(values['заліки'])+current_consultations*float(values['поточні_консультації'])*student_number/float(values['студ_зал'])+student_number*kil_indiv_zavd*float(k_indiv)+float(result_kr)*(float(values['кр'])+float(values['зах_кр']))+float(result_kp)*(float(values['кп'])+float(values['зах_кп']))+float(kil_tij_vir_pr)*float(values['вир_пр_переддипломна'])*student_number+vibir_disc*student_number*float(k_vibirkovi_disc)+kil_atec_ex*student_number*float(values['атест_ЕК'])+findatect(sheetfile1,student_number,values, sheetfile1.title)

        file_save_sheet['Z'+str(row_number)] = toFixed(navantajenya, 2)




        print('navantajenya2 = ', navantajenya)
        npp_bills = npp_counter(values, navantajenya)
        file_save_sheet['AA'+str(row_number)] = toFixed(npp_bills, 2)
        all_bills = bill_counter(values, npp_bills)
        file_save_sheet['AB'+str(row_number)] = toFixed(all_bills, 2)
        bill_of_student = all_bills/student_number
        all_counts.append((navantajenya, npp_bills, all_bills, bill_of_student))
        file_save_sheet['AC'+str(row_number)] = toFixed(all_bills, 2)/student_number

        row_number += 1
    # file_save_sheet.title = sheetfile1.title
    return 'Результати збережені у файл "розрахунок вартості.xlsx!"'


if __name__ == "__main__":
    values = {0: '580', 1: '17100', 2: '1.22', 3: '10590.41', 4: '', 'Переглянути': '', 5: '', 'Переглянути0': '', 'filetrue': False, 6: True, 7: '', 'dop_file': '', 'вибір_дисц_бакалавр_денна': '0.7', 'вибір_дисц_магістр_денна': '0.76', 'вибір_дисц_бакалавр_заочна': '0.17', 'вибір_дисц_магістр_заочна': '0.1', 'вибір_дисц_бакалавр_вечірня': '0.64', 'екз': '0.25', 'пров_екз': '0', 'конс_пред_екз': '2', 'заліки': '0', 'студ_зал': '25', 'пот_конс_денна': '2', 'пот_конс_вечірня': '2', 'пот_конс_заочна': '4', 'пот_конс_дуальна': '10', 'академ_груп': '0.04', 'індивід_денна/вечірня': '0', 'індивід_заочна': '0', 'індивід_дуальна': '0', 'кр': '2', 'кп': '3', 'зах_кр': '1', 'зах_кп': '1', 'нав_практ1': '20', 'нав_практ2': '1', 'вир_практ1': '0', 'вир_практ2': '0.5', 'вир_пр_переддипломна': '0.5', 'студ_нав_практ': '15', 'студ_вир_практ': '90', 'атест_ЕК': '2', 'атест_екз_консультації': '8', 'квал_роб_керівництво1': '0.5', 'квал_роб_керівництво2_до_5к': '3', 'квал_роб_керівництво2_5_та_6к': '10.5', 'квал_роб_рецензування_до_5к': '0', 'квал_роб_рецензування_5_та_6к': '0', 'поток': '90', 'D1coeff': '30', 'E1coeff': '15', 'НПП_витрати': '48', 'поточні_консультації': '0.02', 8: 'D:/Учеба/7_семестр/курсавая/tests/БГ 2022.xlsx', 'plan_vart': 'D:/Учеба/7_семестр/курсавая/tests/БГ 2022.xlsx', 'НПП_витрати1': '48', 'бюджет': '70000', 'контракт': '35000', 9: 'Вартість'}
    print(count_price_file_open(values['plan_vart'], values=values))
    # for i in range(1,61):
    #     print(i, amount_of_subgroups_counter(i, 15))
